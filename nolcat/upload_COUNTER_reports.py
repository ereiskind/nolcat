import logging
import re
from datetime import date
from datetime import datetime
import html
from random import choice
from openpyxl import load_workbook
import pandas as pd
from pandas.api.types import is_string_dtype

from .app import *
from .models import *

log = logging.getLogger(__name__)


class UploadCOUNTERReports:
    """A class for transforming uploaded Excel workbook(s) with tabular COUNTER data for loading into the `COUNTERData` relation.

    COUNTER reports not delivered by SUSHI are given in a tabular format and usually saved in Excel workbooks. These workbooks can be ingested into this program via a Flask-WTF MultipleFileField form field, but that workbook data requires manipulation and cleaning to become a single dataframe that can be loaded into the `COUNTERData` relation. This class exists to make those changes; since the desired behavior is more that of a function than a class, the would-be function becomes a class by dividing it into the traditional `__init__` method, which instantiates the list of Werkzeug FileStorage object(s), each of which encapsulates a selected Excel workbook, as a class attribute, and the `create_dataframe()` method, which performs the actual transformation. This structure requires all instances of the class constructor to be prepended to a call to the `create_dataframe()` method, which means objects of the `UploadCOUNTERReports` type are never instantiated.

    Attributes:
        self.COUNTER_report_files (list): The constructor method for `UploadCOUNTERReports`, which instantiates the list of werkzeug.datastructures.FileStorage objects containing the COUNTER reports to be uploaded.

    Methods:
        create_dataframe: This method transforms the data from the tabular COUNTER reports in uploaded Excel workbooks into a single dataframe ready for normalization.
    """
    def __init__(self, COUNTER_report_files):
        """The constructor method for `UploadCOUNTERReports`, which instantiates the list of werkzeug.datastructures.FileStorage objects containing the COUNTER reports to be uploaded.

        This constructor is not meant to be used alone; all class instantiations should have a `create_dataframe()` method call appended to it.

        Args:
            COUNTER_report_files (list): The list of Werkzeug FileStorage object(s), each of which encapsulates a single uploaded Excel workbook of tabular COUNTER data
        """
        self.COUNTER_report_files = COUNTER_report_files
    

    def create_dataframe(self):
        """This method transforms the data from the tabular COUNTER reports in uploaded Excel workbooks into a single dataframe ready for normalization.

        This method prepares the data from tabular COUNTER reports for upload into the database. This method transforms tabular COUNTER reports on sheets in Excel workbooks into dataframes, but to become complete and valid records in the relation, enhancements are needed, including cleaning the data, filling in data R4 provided through its multiple different report types, and adding the statistics source of the data, which is taken from the first part of the file name.

        Returns:
            tuple: COUNTER data ready for normalization (dataframe); list of workbooks and worksheets with data not in dataframe
        """
        '''Known issues with specific stats sources (taken from webpage instructions):
            * Gale reports needed to be copied and pasted as values with the paste special dialog box to work in OpenRefine
            * iG Press/BEP reports have multiple ISBNs and ISSNs in the fields for those values
        '''
        log.info("Starting `UploadCOUNTERReports.create_dataframe()`.")
        all_dataframes_to_concatenate = []
        data_not_in_dataframes = []
        valid_report_types = ("BR1", "BR2", "BR3", "BR5", "DB1", "DB2", "JR1", "JR2", "MR1", "PR1", "TR1", "TR2", "PR", "DR", "TR", "IR")
        dates_as_string_regex = re.compile(r"([A-Z][a-z]{2})\-(\d{4})")


        #Section: Load the Workbook(s)
        for FileStorage_object in self.COUNTER_report_files:
            log.debug(f"Starting iteration for uploading workbook {FileStorage_object}.")
            # When using the web app, `FileStorage_object` is <class 'werkzeug.datastructures.FileStorage'>; `FileStorage_object.stream._file` is <class '_io.BytesIO'>
            try:
                file = load_workbook(filename=FileStorage_object.stream._file, read_only=True)
                log.debug(f"Successfully loaded the workbook {str(FileStorage_object.filename)}.")
            except Exception as error:
                log.error(f"Loading the workbook {str(FileStorage_object.filename)} raised the error {error}.")
                data_not_in_dataframes.append(f"Workbook {str(FileStorage_object.filename)}")
                continue
            
            try:
                statistics_source_ID = int(re.search(r"(\d+)_.+\.xlsx", str(FileStorage_object.filename)).group(1))
            except Exception as error:
                log.warning(f"The workbook {str(FileStorage_object.filename)} wasn't be loaded because attempting to extract the statistics source ID from the file name raised {error}. Remember the program is looking for a file with a name that begins with the statistics source ID followed by an underscore and ends with the Excel file extension.")
                data_not_in_dataframes.append(f"Workbook {str(FileStorage_object.filename)}")
                continue

            for report_type in file.sheetnames:
                if report_type not in valid_report_types:
                    log.warning(f"The sheet name {report_type} isn't a valid report type, so the sheet couldn't be loaded. Please correct the sheet name and try again.")
                    data_not_in_dataframes.append(f"Worksheet {report_type} in workbook {str(FileStorage_object.filename)}")
                    continue
                sheet = file[report_type]  # `report_type` is the name of the sheet as a string, so it can be used as an index operator
                log.info(f"Loading data from sheet {report_type} from workbook {str(FileStorage_object.filename)}.")


                #Section: Identify the Header Row
                # To handle both R4 and R5 reports, as well as noncompliance with the standard in regards to empty rows in and after the header, the header row of the table is searched for instead of using any presets
                looking_for_header_row = True
                header_row_number = 1

                while looking_for_header_row:
                    count_of_month_labels = 0
                    for cell in sheet[header_row_number]:
                        if cell.value is None or isinstance(cell.value, int) or isinstance(cell.value, float):
                            continue  # `None` and integers/floats (which appear in the "Release" field of the header) cause `TypeError` in `re.fullmatch`, so they need to be weeded out here
                        elif isinstance(cell.value, datetime) or dates_as_string_regex.fullmatch(cell.value) is not None:
                            count_of_month_labels += 1
                    if count_of_month_labels > 1:  # This stops at the first row with multiple dates, which won't be true of any header row
                        number_of_fields = len(sheet[header_row_number])
                        log.debug(f"The table's header is at row {header_row_number}.")
                        looking_for_header_row = False
                        break
                    else:
                        header_row_number += 1
                

                #Section: Get the Field Names
                df_field_names = []
                df_date_field_names = []
                for iterable_of_field_names in sheet.iter_rows(
                    min_row=header_row_number,
                    max_row=header_row_number,
                    min_col=1,
                    max_col=number_of_fields,
                    values_only=True,
                ):  # Creates a tuple with the field names as elements
                    for field_name in iterable_of_field_names:
                        log.debug(f"Getting standardized field name for field {field_name} (type {type(field_name)}).")

                        # `None` in regex methods raises a TypeError, so they need to be in try-except blocks
                        try:
                            if re.fullmatch(r"[Cc]omponent", field_name):
                                continue  # The rarely used `Component` subtype fields aren't captured by this program
                        except TypeError:
                            pass

                        try:  
                            date_as_string = dates_as_string_regex.findall(field_name)
                        except TypeError:
                            date_as_string = False
                        
                        if field_name == "ISSN" and (report_type == 'BR1' or report_type == 'BR2' or report_type == 'BR3' or report_type == 'BR5'):
                            df_field_names.append("online_ISSN")  # This is the first name replacement because assigning a certain type of ISSN changes the meaning slightly
                        
                        elif date_as_string:
                            date_tuple = date_as_string[0]
                            if date_tuple[0] == "Jan":
                                month_int = 1
                            elif date_tuple[0] == "Feb":
                                month_int = 2
                            elif date_tuple[0] == "Mar":
                                month_int = 3
                            elif date_tuple[0] == "Apr":
                                month_int = 4
                            elif date_tuple[0] == "May":
                                month_int = 5
                            elif date_tuple[0] == "Jun":
                                month_int = 6
                            elif date_tuple[0] == "Jul":
                                month_int = 7
                            elif date_tuple[0] == "Aug":
                                month_int = 8
                            elif date_tuple[0] == "Sep":
                                month_int = 9
                            elif date_tuple[0] == "Oct":
                                month_int = 10
                            elif date_tuple[0] == "Nov":
                                month_int = 11
                            elif date_tuple[0] == "Dec":
                                month_int = 12
                            df_field_names.append(date(int(date_tuple[1]), month_int, 1))
                            df_date_field_names.append(date(int(date_tuple[1]), month_int, 1))
                        elif isinstance(field_name, datetime):
                            df_field_names.append(date(field_name.year, field_name.month, 1))  # This both ensures the date is the first of the month and removes the unneeded time data
                            df_date_field_names.append(date(field_name.year, field_name.month, 1))
                        
                        elif (field_name is None or field_name == "" or field_name == " ") and (report_type == 'BR1' or report_type == 'BR2' or report_type == 'BR3' or report_type == 'BR5'):
                            df_field_names.append("resource_name")
                        elif field_name == "Collection" and report_type == 'MR1':
                            df_field_names.append("resource_name")
                        elif field_name == "Database" and (report_type == 'DB1' or report_type == 'DB2'):
                            df_field_names.append("resource_name")
                        
                        elif field_name == "Journal" and (report_type == 'JR1' or report_type == 'JR2'):
                            df_field_names.append("resource_name")
                        elif field_name == "Title" and (report_type == 'BR1' or report_type == 'BR2' or report_type == 'BR3' or report_type == 'BR5' or report_type == 'TR1' or report_type == 'TR2'):
                            df_field_names.append("resource_name")
                        elif field_name == "Database" and report_type == 'DR':
                            df_field_names.append("resource_name")
                        elif field_name == "Title" and report_type == 'TR':
                            df_field_names.append("resource_name")
                        elif field_name == "Item" and report_type == 'IR':
                            df_field_names.append("resource_name")
                        
                        elif field_name == "Content Provider" and report_type == 'MR1':
                            df_field_names.append("publisher")
                        elif field_name == "User Activity" and (report_type == 'BR5' or report_type == 'DB1'or report_type == 'PR1'):
                            df_field_names.append("metric_type")
                        elif (field_name == "Access Denied Category" or field_name == "Access denied category") and (report_type == 'BR3' or report_type == 'DB2' or report_type == 'JR2' or report_type == 'TR2'):
                            df_field_names.append("metric_type")
                        
                        # These change field names from R4 and R5 reports to use R5 naming conventions with lowercase letters, so they don't need to be limited to certain reports
                        elif field_name == "Proprietary Identifier" or field_name == "Proprietary_ID":
                            df_field_names.append("proprietary_ID")
                        elif field_name == "Book DOI" or field_name == "Journal DOI" or field_name == "Title DOI":
                            df_field_names.append("DOI")
                        elif field_name == "Data type" or field_name == "Data Type" or field_name == "Data_Type":
                            df_field_names.append("data_type")
                        elif field_name == "Online ISSN" or field_name == "Online_ISSN":
                            df_field_names.append("online_ISSN")
                        elif field_name == "Print ISSN" or field_name == "Print_ISSN":
                            df_field_names.append("print_ISSN")

                        elif field_name is None:
                            continue  # Deleted data and merged cells for header values can make Excel think null columns are in use; when read, these columns add `None` to `df_field_names`, causing a `ValueError: Number of passed names did not patch number of header fields in the file` when reading the worksheet contents into a dataframe
                        
                        elif re.search(r"_((ID)|(DOI)|(URI)|(IS[SB]N))$", field_name):  # The regex captures strings ending with `ID`, `DOI`, `URI`, `ISSN`, and `ISBN` after an underscore; no try-except block is needed because `None` values were filtered out above
                            df_field_names.append("_".join(field_name.split("_")[0:-1]).lower() + "_" + field_name.split("_")[-1])
                        elif field_name == "DOI" or field_name == "URI" or field_name == "YOP" or field_name == "ISBN":  # These field names are just capital letters and should remain that way, so they must be handled separately
                            df_field_names.append(field_name)
                        
                        else:
                            df_field_names.append(field_name.lower())
                df_non_date_field_names = [field_name for field_name in df_field_names if field_name not in df_date_field_names]  # List comprehension used to preserve order
                log.info(f"The COUNTER report contains the fields {df_non_date_field_names} and data for the dates {df_date_field_names}.")


                #Section: Create Dataframe
                #Subsection: Create Data Type Dictionary
                df_dtypes = {k: v for (k, v) in COUNTERData.state_data_types().items() if k in df_field_names}
                date_dtype_fields = {k: v for (k, v) in df_dtypes.items() if v == "datetime64[ns]"}

                #Subsection: Create Dataframe from Excel Worksheet
                df = pd.read_excel(
                    FileStorage_object.stream._file,  # The method accepts the string and BytesIO objects this variable provides when using the test module and the web app respectively
                    sheet_name=report_type,
                    engine='openpyxl',
                    header=header_row_number-1,  # This gives the row number with the headings in Excel, which is also the row above where the data starts
                    names=df_field_names,
                    dtype={k: v for (k, v) in df_dtypes.items() if v != "datetime64[ns]"},  # Ensuring string fields are set as such keeps individual values within those fields from being set as numbers or dates (e.g. resources with a date or year for a title)
                )
                log.info(f"Dataframe immediately after creation:\n{df.head()}\n{return_string_of_dataframe_info(df)}")
                log.debug(f"Complete dataframe:\n{df}")
                for field in date_dtype_fields.keys():
                    df[field] = pd.to_datetime(
                        df[field],
                        errors='coerce',  # Changes the null values to the date dtype's null value `NaT`
                        utc=True,  # This must be set to `True` to convert timezone-aware datetime objects
                    )
                    df[field] = df[field].dt.tz_localize(None)
                log.debug(f"Dataframe after all initial dtypes set\n{return_string_of_dataframe_info(df)}")


                #Section: Make Pre-Stacking Updates
                df = df.replace(r"\n", "", regex=True)  # Removes errant newlines found in some reports, primarily at the end of resource names
                df = df.map(lambda cell_value: html.unescape(cell_value) if isinstance(cell_value, str) else cell_value)  # Reverts all HTML escaped values

                #Subsection: Make Publication Dates Date Only ISO Strings
                # At this point, dates can be in multiple formats and data types, but NoLCAT doesn't need time or timezone data, and since all the metadata values are combined into a larger string as part of the unstacking process, using ISO strings or the null placeholder string is appropriate
                if "publication_date" in df_field_names:
                    df['publication_date'] = df['publication_date'].fillna("`None`")  # Different data types use different null values, so switching to the null placeholder string now prevents type juggling issues
                    df['publication_date'] = df['publication_date'].apply(lambda cell_value: str(cell_value).split("T")[0] if isinstance(cell_value, str) else cell_value)
                    df['publication_date'] = df['publication_date'].apply(lambda cell_value: cell_value.strftime('%Y-%m-%d') if isinstance(cell_value, datetime) else cell_value)  # Date data types in pandas inherit from `datetime.datetime`
                    df['publication_date'] = df['publication_date'].apply(lambda cell_value: "`None`" if cell_value=='1000-01-01' or cell_value=='1753-01-01' or cell_value=='1900-01-01' else cell_value)
                if "parent_publication_date" in df_field_names:
                    df['parent_publication_date'] = df['parent_publication_date'].fillna("`None`")  # Different data types use different null values, so switching to the null placeholder string now prevents type juggling issues
                    df['parent_publication_date'] = df['parent_publication_date'].apply(lambda cell_value: str(cell_value).split("T")[0] if isinstance(cell_value, str) else cell_value)
                    df['parent_publication_date'] = df['parent_publication_date'].apply(lambda cell_value: cell_value.strftime('%Y-%m-%d') if isinstance(cell_value, datetime) else cell_value)  # Date data types in pandas inherit from `datetime.datetime`
                    df['parent_publication_date'] = df['parent_publication_date'].apply(lambda cell_value: "`None`" if cell_value=='1000-01-01' or cell_value=='1753-01-01' or cell_value=='1900-01-01' else cell_value)

                #Subsection: Add `statistics_source_ID` and `report_type` Fields
                # The names of the fields are added at the same time as the fields themselves so the length of the list of field names matches the number of fields in the dataframe
                df['statistics_source_ID'] = statistics_source_ID
                df_field_names.append("statistics_source_ID")
                df_non_date_field_names.append("statistics_source_ID")

                df['report_type'] = report_type
                df_field_names.append("report_type")
                df_non_date_field_names.append("report_type")
                log.debug(f"Dataframe field names: {df_field_names}")

                #Subsection: Remove `Reporting Period` Field
                df_field_names_sans_reporting_period_fields = [field_name for field_name in df_non_date_field_names if not re.search(r"[Rr]eporting[\s_][Pp]eriod", field_name)]
                reporting_period_field_names = [field_name for field_name in df_non_date_field_names if field_name not in df_field_names_sans_reporting_period_fields]  # List comprehension used to preserve list order
                df = df.drop(columns=reporting_period_field_names)
                df_field_names = df_field_names_sans_reporting_period_fields + df_date_field_names
                log.debug(f"Dataframe field names with statistics source ID and without reporting period: {df_field_names}")

                #Subsection: Remove Total Rows
                if re.fullmatch(r"PR1?", report_type) is None:
                    log.debug("About to remove total rows from non-platform reports.")
                    number_of_rows_with_totals = df.shape[0]
                    common_summary_rows = df['resource_name'].str.contains(r"^[Tt]otal\s[Ff]or\s[Aa]ll\s\w+", regex=True)  # `\w+` is because values besides `title` are used in various reports
                    uncommon_summary_rows = df['resource_name'].str.contains(r"^[Tt]otal\s[Ss]earches", regex=True)
                    summary_rows = common_summary_rows | uncommon_summary_rows
                    summary_rows.name = 'summary_rows'  # Before this, the series is named `resource_name`, just like the series it was filtered from
                    df = df.join(summary_rows)
                    df = df[~df['summary_rows']]
                    df = df.drop(columns=['summary_rows'])
                    log.debug(f"Number of rows in report of type {report_type} reduced from {number_of_rows_with_totals} to {df.shape[0]}.")

                #Subsection: Split ISBNs and ISSNs in TR
                if re.fullmatch(r"TR[1|2]", report_type):
                    log.debug("About to separate identifiers in COUNTER R4 title report.")
                    # Creates fields containing `True` if the original field's value matches the regex, `False` if it doesn't match the regex, and null if the original field is also null
                    df['print_ISSN'] = df['Print ID'].str.match(ISSN_regex())
                    df['online_ISSN'] = df['Online ID'].str.match(ISSN_regex())
                    # Returns `True` if the values of `print_ISSN` and `online_ISSN` are `True`, otherwise, returns `False`
                    df['ISBN'] = df['print_ISSN'] & df['online_ISSN']

                    # Replaces Booleans signaling value with values from `Print ID` and `Online ID`
                    df.loc[df['print_ISSN'] == True, 'print_ISSN'] = df['Print ID']
                    df.loc[df['online_ISSN'] == True, 'online_ISSN'] = df['Online ID']
                    df.loc[df['print_ISSN'] == False, 'ISBN'] = df['Print ID']
                    df.loc[df['online_ISSN'] == False, 'ISBN'] = df['Online ID']
                    # Replace Booleans not signaling value with null placeholder string (replacing with `None` causes the values to fill down instead of the desired replacement)
                    df['print_ISSN'] = df['print_ISSN'].replace(False, "`None`")  
                    df['online_ISSN'] = df['online_ISSN'].replace(False, "`None`")
                    df['ISBN'] = df['ISBN'].replace(True, "`None`")
                    df['ISBN'] = df['ISBN'].replace(False, "`None`")  # These cells occur when the ID columns have a null value and an ISSN

                    # Update field names and order in `df_field_names` and dataframe
                    df = df.drop(columns=['Print ID', 'Online ID'])
                    df_field_names.remove("Print ID")
                    df_field_names.remove("Online ID")
                    df_field_names.insert(len(df_field_names)-len(df_date_field_names)-1, "ISBN")
                    df_field_names.insert(len(df_field_names)-len(df_date_field_names)-1, "print_ISSN")
                    df_field_names.insert(len(df_field_names)-len(df_date_field_names)-1, "online_ISSN")
                    df = df[df_field_names]
                    log.debug(f"Dataframe with identifiers in standardized fields:\n{df}")

                #Subsection: Put Placeholder in for Null Values
                df = df.fillna("`None`")
                log.debug("Null values in dataframe replaced with string placeholder.")
                df = df.replace(
                    to_replace='^\s*$',
                    # The regex is designed to find the blank but not null cells by finding those cells containing nothing (empty strings) or only whitespace. The whitespace metacharacter `\s` is considered invalid by libraries not used in this class, and without the anchors, the replacement is applied not just to whitespaces but to spaces between characters as well.
                    value="`None`",
                    regex=True
                )

                log.debug(f"Dataframe with pre-stacking changes:\n{df}\n{return_string_of_dataframe_info(df)}")


                #Section: Stack Dataframe
                #Subsection: Create Lists for Sorting Fields
                list_of_field_names_from_df = df.columns.values.tolist()
                df_non_date_field_names = [field_name for field_name in df_field_names if field_name not in df_date_field_names]  # Reassigning this variable with the same statement because one of the values in the statement has changed
                boolean_identifying_metadata_fields = [True if field_name in df_non_date_field_names else False for field_name in list_of_field_names_from_df]

                fields_and_their_dtypes = {field_name: field_dtype for (field_name, field_dtype) in COUNTERData.state_data_types().items() if field_name in list_of_field_names_from_df}
                list_of_string_fields = [field_name for (field_name, field_dtype) in fields_and_their_dtypes.items() if field_dtype == "string"]  # Not actually doing a dtype conversion because null values were replaced with a string placeholder

                #Subsection: Determine Delimiter Character
                # To properly separate the values being combined in the next subsection, the delimiter cannot be present in any of the fields being combined, and a single character must be used because pandas 1.3 doesn't seem to handle multi-character literal string delimiters. Possible delimiters are tested before their use to prevent problems later on.
                possible_delimiter_characters = ['#', '~', '@', '^', '`', '|', '$', '*', '†', '‡']  # Hash is the first tested delimiter because it appears in a title in the test data, so this aspect of the code is covered by the tests
                string_type_df_fields = [field_name for field_name in df_non_date_field_names if field_name in list_of_string_fields]
                delimiters_in_data = []
                for character in possible_delimiter_characters:
                    for field in string_type_df_fields:
                        log.debug(f"Checking for delimiter '{character}' in field {field}.")
                        if df[field].apply(lambda cell_value: character in cell_value).any():
                            log.debug(f"Delimiter '{character}' found in field {field}.")
                            delimiters_in_data.append(character)
                            break
                        else:
                            log.debug(f"Delimiter '{character}' not found in field {field}.")
                
                potential_delimiters = [character for character in possible_delimiter_characters if character not in delimiters_in_data]
                if potential_delimiters:
                    delimiter_character = choice(potential_delimiters)
                    log.info(f"Using '{delimiter_character}' as the delimiter.")
                else:
                    message = "None of the possible delimiter characters were viable, so the `delimiter_character` variable, which the program needs to continue, wasn't set."
                    log.critical(message)
                    data_not_in_dataframes.append(f"Worksheet {report_type} in workbook {str(FileStorage_object.filename)}")
                    continue

                #Subsection: Create Temp Index Field with All Metadata Values
                # Combines all values in the fields specified by the index operator of the dataframe to which the `apply` method is applied
                df['temp_index'] = df[df.columns[boolean_identifying_metadata_fields]].apply(
                    lambda cell_value: delimiter_character.join(cell_value.astype(str)),  
                    axis='columns'
                )
                df = df.drop(columns=df_non_date_field_names)
                log.debug(f"Dataframe with without metadata columns:\n{df}\n{return_string_of_dataframe_info(df)}")
                df = df.set_index('temp_index')
                log.debug(f"Dataframe with new index column:\n{df}")

                #Subsection: Reshape with Stacking
                df = df.stack()  # This creates a series with a multiindex: the multiindex is the metadata, then the dates; the data is the usage counts
                log.debug(f"Dataframe immediately after stacking:\n{df}")
                df = df.reset_index(name='usage_count').rename(columns={"level_1": "usage_date"})
                log.debug(f"Dataframe with reset index:\n{df}\n{return_string_of_dataframe_info(df)}")

                #Subsection: Recreate Metadata Fields
                df[df_non_date_field_names] = df['temp_index'].str.split(pat=delimiter_character, expand=True)  # This splits the metadata values in the index at the chosen delimiter into their own fields and applies the appropriate names to those fields
                log.debug(f"Dataframe after splitting temp index:\n{return_string_of_dataframe_info(df)}")
                df = df.drop(columns='temp_index')
                log.debug(f"Fully transposed dataframe:\n{df}")


                #Section: Adjust Data in Dataframe
                #Subsection: Remove Rows with No Usage
                df = df[df['usage_count'] != 0]
                df = df[df['usage_count'] != "`None`"]  # Some zero `usage_count` values may have been null values replaced by the null placeholder; retaining them interferes with correcting data types
                df = df.reset_index(drop=True)
                log.debug(f"Dataframe with zero usage records removed:\n{df}")

                #Subsection: Correct Data Types, Including Replacing Null Placeholders with Null Values
                if "YOP" in df_dtypes.keys():
                    df_dtypes['YOP'] = df_dtypes['YOP'].lower()  # The `YOP` field cannot be converted directly to a pandas nullable int type; this overwrites that dtype value from the `COUNTERData.state_data_types()` method in favor of an intermediary numpy dtype
                log.debug(f"Dataframe info before any null or dtype adjustments:\n{return_string_of_dataframe_info(df)}")
                for field in {k: v for (k, v) in df_dtypes.items() if v != "string"}.keys():  # The null placeholders need to be converted in non-string fields before the dtype conversion because the placeholders are strings and thus can't be converted into the other types
                    if field == "YOP":  # At this point, is `int16`, which doesn't accept null values
                        df[field] = df[field].replace(["`None`"], [0])  # Values must be enclosed in lists for method to work
                    else:
                        df[field] = df[field].replace(["`None`"], [None])  # Values must be enclosed in lists for method to work
                    log.debug(f"Dataframe info after removing null placeholders in `{field}`:\n{return_string_of_dataframe_info(df)}")
                log.debug(f"Dataframe info before dtype conversion:\n{return_string_of_dataframe_info(df)}")
                if "YOP" in df_dtypes.keys():
                    df['YOP'] = df['YOP'].apply(lambda cell_value: str(cell_value)[:-2] if str(cell_value).endswith(".0") else cell_value)
                df = df.astype(df_dtypes)
                if "YOP" in df_dtypes.keys():
                    df = df.astype({'YOP': COUNTERData.state_data_types()['YOP']})  # This converts the `YOP` field from the intermediary numpy dtype to the final pandas dtype
                    df['YOP'] = df['YOP'].replace(0, pd.NA)
                log.debug(f"Dataframe info after dtype conversion:\n{return_string_of_dataframe_info(df)}")
                df = df.replace(["`None`"], [None])  # The null placeholders need to be converted in string fields after the dtype conversion because having `NoneType` values in fields can cause object to string conversion to fail
                log.debug(f"Dataframe info after dtype and null conversions:\n{return_string_of_dataframe_info(df)}")

                #Subsection: Add Fields Missing from R4 Reports
                if report_type == 'BR1' or report_type == 'BR2' or report_type == 'BR3' or report_type == 'BR5':
                    df['data_type'] = "Book"
                elif report_type == 'DB1' or report_type == 'DB2':
                    df['data_type'] = "Database"
                elif report_type == 'JR1' or report_type == 'JR2':
                    df['data_type'] = "Journal"
                elif report_type == 'MR1':
                    df['data_type'] = "Multimedia"
                elif report_type == 'PR1':
                    df['data_type'] = "Platform"

                if report_type == 'BR1' or report_type == 'BR3' or report_type == 'BR5':
                    df['section_type'] = "Book"
                elif report_type =='BR2':
                    df['section_type'] = "Book_Segment"
                elif report_type == 'JR1' or report_type == 'JR2':
                    df['section_type'] = "Article"
                elif report_type == 'TR1' or report_type == 'TR2':
                    df.loc[df['data_type'] == "Journal", 'section_type'] = "Article"
                    df.loc[df['data_type'] == "Book", 'section_type'] = "Book_Segment"
                    # Any data types besides `Book` or `Journal` won't have a section type

                if report_type =='BR1' or report_type == 'TR1':
                    df['metric_type'] = "Successful Title Requests"
                elif report_type =='BR2':
                    df['metric_type'] = "Successful Section Requests"
                elif report_type =='JR1':
                    df['metric_type'] = "Successful Full-text Article Requests"
                elif report_type =='MR1':
                    df['metric_type'] = "Successful Content Unit Requests"
                
                #Subsection: Standardize R4 Turnaway Metric Values
                df['metric_type'] = df['metric_type'].apply(lambda cell_value: cell_value.replace("licence", "license"))  #  Always use American English spelling for `license`
                df['metric_type'] = df['metric_type'].apply(lambda cell_value: cell_value.replace("denied. C", "denied: c"))
                df['metric_type'] = df['metric_type'].apply(lambda cell_value: cell_value.replace("denied.", "denied:"))
                df['metric_type'] = df['metric_type'].apply(lambda cell_value: cell_value.replace("denied: C", "denied: c"))
                
                log.info(f"Dataframe being used in concatenation:\n{df.head()}\n{return_string_of_dataframe_info(df)}")
                log.debug(f"Complete dataframe:\n{df}")
                all_dataframes_to_concatenate.append(df)
        
        
        #Section: Create Final Combined Dataframe
        #Subsection: Combine Dataframes
        combined_df = pd.concat(
            all_dataframes_to_concatenate,
            ignore_index=True,  # Resets index
        )
        log.info(f"Combined dataframe info:\n{return_string_of_dataframe_info(combined_df)}")
        log.debug(f"Combined dataframe:\n{combined_df}")

        #Subsection: Set Data Types
        combined_df_field_names = combined_df.columns.values.tolist()
        combined_df = combined_df.astype(
            {k: v for (k, v) in COUNTERData.state_data_types().items() if k in combined_df_field_names},
            errors='ignore',
        )
        if "publication_date" in combined_df_field_names:
            combined_df['publication_date'] = pd.to_datetime(
                combined_df['publication_date'],
                errors='coerce',  # Changes the null values to the date dtype's null value `NaT`
            )
        if "parent_publication_date" in combined_df_field_names:
            combined_df['parent_publication_date'] = pd.to_datetime(
                combined_df['parent_publication_date'],
                errors='coerce',  # Changes the null values to the date dtype's null value `NaT`
            )
        #combined_df = combined_df.fillna(value=None)  # Replacing the pandas and numpy specialized null values with the standard Python null value


        #Section: Return Dataframe
        log.info(f"Final dataframe:\n{combined_df}\nand dtypes:\n{combined_df.dtypes}")
        return (combined_df, data_not_in_dataframes)