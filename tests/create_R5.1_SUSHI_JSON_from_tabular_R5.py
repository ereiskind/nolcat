"""Transform Excel worksheet produced by "tests\\data\\create_JSON_base.json" into a COUNTER R5.1 SUSHI JSON. This module requires pandas 1.4 or higher to run due to the bug described at https://github.com/pandas-dev/pandas/pull/43949."""

import logging
import pathlib
import io
import re
import json
from openpyxl import load_workbook
import pandas as pd


# `from ..nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat import app` raises `ModuleNotFoundError: No module named 'nolcat'`
# `from .. import nolcat` raises `ImportError: attempted relative import with no known parent package`
# `from .nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat.app import return_string_of_dataframe_info` raises `ModuleNotFoundError: No module named 'nolcat'`

logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s] %(message)s")

absolute_path_to_tests_directory = pathlib.Path(__file__).parent.resolve()
directory_with_final_JSONs = absolute_path_to_tests_directory / 'data' / 'R5.1_COUNTER_JSONs_for_tests'


def return_string_of_dataframe_info(df):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    in_memory_stream = io.StringIO()
    df.info(buf=in_memory_stream)
    return in_memory_stream.getvalue()


def last_day_of_month(first_day_of_month):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    year_and_month_string = first_day_of_month.date().isoformat()[0:-2]  # Returns an ISO date string, then takes off the last two digits
    return year_and_month_string + str(first_day_of_month.days_in_month)


def remove_null_counts(JSON):
    """This function removes the date and count key-value pairs when the count value is null.

    Part of the dataframe manipulation used to turn the tabular data into a JSON matching the `Performance` section of a SUSHI 5.1 JSON is a pivot, which automatically places null values in cells that don't have any corresponding data values. These are retained through the `to_dict()` method, placing them in the JSON, but since null values aren't included in any COUNTER 5.x reports, they need to be removed.

    Args:
        JSON (dict): the `Performance` section of a SUSHI 5.1 report for a given metadata combination
    
    Returns:
        dict: the argument value sans the date and count key-value pairs where the value is null
    """
    outer_dict = dict()
    for metric_type, date_and_count in JSON.items():
        inner_dict = dict()
        for date, count in date_and_count.items():
            if pd.notnull(count):
                inner_dict[date] = count
        outer_dict[metric_type] = inner_dict
    return outer_dict


def update_method_with_return_value(dict1, dict2):
    """This function performs the Python `dictionary.update()` method and returns the modified dictionary.

    The Python `dictionary.update()` method modifies the dictionary in place and returns `None`; this function is a way to get the result of that method as a return value.

    Args:
        dict1 (dict): a dictionary running the `update()` method
        dict2 (dict): a dictionary fed into the `update()` method

    Returns:
        dict: the combination of all the key-value pairs in `dict1` and `dict2`
    """
    dict1.update(dict2)
    return dict1


#Section: Load the Workbook(s)
file_path = input("Enter the complete file path for the Excel workbook output from OpenRefine: ")
file_path = pathlib.Path(file_path)
file = load_workbook(filename=file_path, read_only=True)
report_name = file_path.stem
report_type = report_name.split("_")[1]
sheet = file[report_name]
logging.info(f"Creating JSON from report {report_name} for a {report_type}.")


#Section: Get the Field Names
# Excel workbooks created by Apache POI programs like OpenRefine don't have their `max_column` and `max_row` attributes set properly, so those values must be reset, after which a loop through rows that breaks after the first row can be used
sheet.reset_dimensions()
df_field_names = []
for row in sheet.rows:
    for field_name in row:
        
        # `None` in regex methods raises a TypeError, so they need to be in try-except blocks
        try:
            if re.match(r'^[Cc]omponent', field_name.value):
                continue  # The rarely used `Component` subtype fields aren't captured by this program
        except TypeError:
            pass
        
        if field_name.value is None:
            continue  # Deleted data and merged cells for header values can make Excel think null columns are in use; when read, these columns add `None` to `df_field_names`, causing a `ValueError: Number of passed names did not patch number of header fields in the file` when reading the worksheet contents into a dataframe
        else:
            df_field_names.append(field_name.value)
    break
logging.info(f"The field names are {df_field_names}.")


#Section: Create Dataframe
#Subsection: Ensure String Data Type for Potentially Numeric Metadata Fields
# Strings will be pandas object dtype at this point, but object to string conversion is fairly simple; string fields that pandas might automatically assign a numeric dtype to should be set as strings at the creation of the dataframe to head off problems.
df_dtypes = dict()
for field_name in df_field_names:
    if field_name == "Count" or field_name == "YOP":
        df_dtypes[field_name] = 'Int64'  # The pandas integer dtype; Python's 'int' is ignored
    else:  # JSON uses strings for dates; `Begin_Date` is changed to a date with the `converters` argument in `read_excel`
        df_dtypes[field_name] = 'string'

#Subsection: Create Dataframe from Excel Worksheet
df = pd.read_excel(
    file_path,
    sheet_name=report_name,
    engine='openpyxl',
    header=0,  # This means the first row of the spreadsheet will be treated as the row with the headers, meaning the data starts on the second row
    names=df_field_names,
    dtype=df_dtypes,
    converters={  # `to_datetime` called as object, not function; adding `format` argument causes TypeError
    'Begin_Date': pd.to_datetime,
    'Publication_Date': pd.to_datetime,
    'Parent_Publication_Date': pd.to_datetime,
    }
)
logging.debug(f"Complete dataframe:\n{df}")
logging.info(f"Dataframe summary info:\n{return_string_of_dataframe_info(df)}")
number = 0  # The output often needs more investigation that can be provided in the logging stdout, so frequent `to_csv()` and `to_json()` calls are made; this variable gives the order of the calls


#Section: Update Dataframe
df = df.replace(r'\n', '', regex=True)  # Removes errant newlines found in some reports, primarily at the end of resource names
df = df.replace("licence", "license")  # "Have `license` always use American English spelling
try:
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')
except:
    df['Begin_Date'] = pd.to_datetime(df['Begin_Date'])
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')

#Subsection: Put Placeholder in for Null Values
# Null values and fields with nothing but null values cause problems and errors in groupby functions, so they're replaced with placeholder strings
df = df.fillna("`None`")
df = df.replace(
    to_replace='^\s*$',
    # The regex is designed to find the blank but not null cells by finding those cells containing nothing (empty strings) or only whitespace. The whitespace metacharacter `\s` is marked with a deprecation warning, and without the anchors, the replacement is applied not just to whitespaces but to spaces between characters as well.
    value="`None`",
    regex=True
)
logging.debug(f"Dataframe after initial updates:\n{df}")
####################
#output = df.copy()
#purpose = "after-initial-updates"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################


######Section: CHANGE DATAFRAME INTO JSON #####

#Section: Create Field Lists and Dataframes for Multiindexes, Groupby Operations, and Dataframe Recombination
fields_in_performance = ['Begin_Date', 'Metric_Type', 'Count']
metadata_multiindex_fields = [field_name for field_name in df_field_names if field_name not in fields_in_performance]

fields_used_in_join_multiindex = metadata_multiindex_fields + ['Begin_Date']
join_multiindex_df = df[fields_used_in_join_multiindex].set_index(metadata_multiindex_fields, drop=False)
####################
#output = join_multiindex_df.copy()
#purpose = "multiindex-for-joining"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
groupby_multiindex = [f"index_{field_name}" for field_name in metadata_multiindex_fields]  # All sections have `index_` appended to the beginning of the index field names before the index reset for record deduplication, so `index_` needs to be at the beginning of the fields in the list in the groupby operation
fields_to_drop_at_end = []

# The default text sorting methods in OpenRefine and Pandas order non-letter characters differently, meaning applying a pandas sort to the resource names won't put them in the same order as they were in openRefine after the sort; to return the records to their same order at the end of the module, their order must be saved at this point
record_sorting_strings = df[metadata_multiindex_fields].apply(
    lambda cell_value: '|'.join(cell_value.astype(str)),  # Combines all values in the fields specified by the index operator of the dataframe to which the `apply` method is applied
    axis='columns',
)
record_sorting_strings = record_sorting_strings.drop_duplicates().reset_index(drop=True)
logging.debug(f"Complete metadata in order from OpenRefine:\n{record_sorting_strings}")
record_sorting_dict = record_sorting_strings.to_dict()
record_sorting_dict = {metadata_string: order_number for (order_number, metadata_string) in record_sorting_dict.items()}
logging.info(f"Metadata strings with ordering numbers:\n{record_sorting_dict}")

#Subsection: Create `Attribute_Performance` Metadata Field Lists
fields_outside_attribute_performance = [
    "Platform",  # PR, DR
    "Database",  # DR
    "Publisher",  # DR
    "Publisher_ID",  # DR
    "Proprietary_ID", # DR
]
metadata_outside_attribute_performance = [field for field in df_field_names if field in fields_outside_attribute_performance]

fields_inside_attribute_performance = [
    "Data_Type",  # PR, DR
    "Access_Method",  # PR, DR
]
metadata_inside_attribute_performance = [field for field in df_field_names if field in fields_inside_attribute_performance]


#Section: Organize Metadata in `Performance`
performance_df = df.copy()
#Subsection: Pivot Data
pivot_index = metadata_multiindex_fields + ['Metric_Type']
performance_df = performance_df.pivot(index=pivot_index, columns='Begin_Date', values='Count')
####################
#output = performance_df.copy()
#purpose = "create-performance-df"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.debug(f"`performance_df` after pivot:\n{performance_df}")

#Subsection: Set Field Names
performance_df = performance_df.reset_index().set_index('Metric_Type')
performance_index_names = {field_name:field_name[:-3] for field_name in performance_df.columns.to_list() if field_name not in metadata_multiindex_fields}
performance_index_names.update({field_name:f"index_{field_name}" for field_name in performance_df.columns.to_list() if field_name in metadata_multiindex_fields})  # The method makes the change in place
performance_df = performance_df.rename(columns=performance_index_names)
####################
#output = performance_df.copy()
#purpose = "performance-df-field-names"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Create JSON Field
performance_df = (performance_df.groupby(groupby_multiindex)).apply(lambda x: x[[field_name for field_name in performance_df.columns.to_list() if field_name not in groupby_multiindex]].to_dict('index')).rename("Performance")
####################
#output = performance_df.copy()
#purpose = "performance-df-JSON-series"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
performance_df = performance_df.apply(lambda performance_JSON: remove_null_counts(performance_JSON))
####################
output = performance_df.copy()
purpose = "final-performance-df"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.debug(f"`performance_df`\n{performance_df}")
logging.debug(f"JSON with `Performance` nesting:\n{performance_df.to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Organize Metadata in `Attribute_Performance`
inside_attribute_performance_df = join_multiindex_df.copy()
inside_attribute_performance_df = inside_attribute_performance_df[metadata_inside_attribute_performance]
####################
#output = inside_attribute_performance_df.copy()
#purpose = "create-inside-attribute-df"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Update Null Values and Index Field Names
inside_attribute_performance_df = inside_attribute_performance_df.replace({"`None`": None})
inside_attribute_performance_df.index = inside_attribute_performance_df.index.set_names({field_name:f"index_{field_name}" for field_name in inside_attribute_performance_df.index.names})
####################
#output = inside_attribute_performance_df.copy()
#purpose = "inside-attribute-df-rename-index"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
inside_attribute_performance_df = inside_attribute_performance_df.reset_index()
####################
#output = inside_attribute_performance_df.copy()
#purpose = "inside-attribute-df-index-reset"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Deduplicate Records
inside_attribute_performance_df['repeat'] = inside_attribute_performance_df.duplicated(keep='first')
inside_attribute_performance_df =  inside_attribute_performance_df.loc[inside_attribute_performance_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
inside_attribute_performance_df =  inside_attribute_performance_df.drop(columns=['repeat'])
####################
#output = inside_attribute_performance_df.copy()
#purpose = "inside-attribute-df-remove-duplicate-records"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#ToDo: Create other nested subsections

#Subsection: Add `Performance` Section as Key-Value Pair
inside_attribute_performance_df = inside_attribute_performance_df.set_index(groupby_multiindex)
inside_attribute_performance_df = pd.concat([inside_attribute_performance_df, performance_df], axis='columns', ignore_index=False)
####################
#output = inside_attribute_performance_df.copy()
#purpose = "inside-attribute-df-add-performance"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Create JSON Field
inside_attribute_performance_df = (inside_attribute_performance_df.groupby(groupby_multiindex)).apply(lambda inside_groupby_df: inside_groupby_df[metadata_inside_attribute_performance + ['Performance']].to_dict('records')[0]).rename("temp_Attribute_Performance")
####################
output = inside_attribute_performance_df.copy()
purpose = "final-inside-attribute-df"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.debug(f"`inside_attribute_performance_df`\n{inside_attribute_performance_df}")
logging.debug(f"JSON with `Attribute_Performance` nesting:\n{inside_attribute_performance_df.to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Organize Metadata Outside `Attribute_Performance`
outside_attribute_performance_df = join_multiindex_df.copy()
outside_attribute_performance_df = outside_attribute_performance_df[metadata_outside_attribute_performance]
####################
#output = outside_attribute_performance_df.copy()
#purpose = "create-outside-attribute-df"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Update Null Values and Index Field Names
outside_attribute_performance_df = outside_attribute_performance_df.replace({"`None`": None})
outside_attribute_performance_df.index = outside_attribute_performance_df.index.set_names({field_name:f"index_{field_name}" for field_name in outside_attribute_performance_df.index.names})
####################
#output = outside_attribute_performance_df.copy()
#purpose = "outside-attribute-df-rename-index"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
outside_attribute_performance_df = outside_attribute_performance_df.reset_index()
####################
#output = outside_attribute_performance_df.copy()
#purpose = "outside-attribute-df-index-reset"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Deduplicate Records
outside_attribute_performance_df['repeat'] = outside_attribute_performance_df.duplicated(keep='first')
outside_attribute_performance_df =  outside_attribute_performance_df.loc[outside_attribute_performance_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
outside_attribute_performance_df =  outside_attribute_performance_df.drop(columns=['repeat'])
####################
#output = outside_attribute_performance_df.copy()
#purpose = "outside-attribute-df-remove-duplicate-records"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=False)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Organize Metadata in `Item_ID`
if report_type == "DR":
    if 'Proprietary_ID' in list(join_multiindex_df.columns):  # If the proprietary ID field, the only item ID field in DR, exists
        fields_to_drop_at_end.append('Proprietary_ID')
        if not join_multiindex_df['Proprietary_ID'].eq("`None`").all():  # If the proprietary ID field has values
            item_ID_values_df = join_multiindex_df.drop(columns=[field_name for field_name in metadata_multiindex_fields if field_name != "Proprietary_ID"])  # This leaves `Begin_Date` for the sake of the groupby later on
            ####################
            #output = item_ID_values_df.copy()
            #purpose = "create-item-ID-df"
            #number = number + 1
            #output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
            #try:
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            #except ValueError:
            #    new_index_names = {name:f"_index_{name}" for name in output.index.names}
            #    output.index = output.index.set_names(new_index_names)
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            ####################
            item_ID_values_df = item_ID_values_df.replace({"`None`": None})
            item_ID_values_df = item_ID_values_df.rename(columns={"Proprietary_ID": "Proprietary"})
            item_ID_values_df.index = item_ID_values_df.index.set_names({field_name:f"index_{field_name}" for field_name in item_ID_values_df.index.names})
            ####################
            #output = item_ID_values_df.copy()
            #purpose = "create-item-ID-rename-fields"
            #number = number + 1
            #output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
            #try:
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            #except ValueError:
            #    new_index_names = {name:f"_index_{name}" for name in output.index.names}
            #    output.index = output.index.set_names(new_index_names)
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            ####################
            item_ID_values_df = item_ID_values_df.reset_index()
            ####################
            #output = item_ID_values_df.copy()
            #purpose = "create-item-ID-reset-index-before-dedupe"
            #number = number + 1
            #output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
            #try:
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            #except ValueError:
            #    new_index_names = {name:f"_index_{name}" for name in output.index.names}
            #    output.index = output.index.set_names(new_index_names)
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            ####################
            item_ID_values_df['repeat'] = item_ID_values_df.duplicated(subset=groupby_multiindex, keep='first')
            item_ID_values_df = item_ID_values_df.loc[item_ID_values_df['repeat'] == False]
            item_ID_values_df = item_ID_values_df.drop(columns=['repeat'])
            ####################
            #output = item_ID_values_df.copy()
            #purpose = "item-ID-df-deduped"
            #number = number + 1
            #output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
            #try:
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            #except ValueError:
            #    new_index_names = {name:f"_index_{name}" for name in output.index.names}
            #    output.index = output.index.set_names(new_index_names)
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            ####################
            item_ID_values_df = (item_ID_values_df.groupby(groupby_multiindex)).apply(lambda item_ID_groupby: item_ID_groupby[['Proprietary']].to_dict('records')[0]).rename("Item_ID")
            ####################
            #output = item_ID_values_df.copy()
            #purpose = "final-item-ID-df"
            #number = number + 1
            #output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
            #try:
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            #except ValueError:
            #    new_index_names = {name:f"_index_{name}" for name in output.index.names}
            #    output.index = output.index.set_names(new_index_names)
            #    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
            ####################
            logging.debug(f"`item_ID_values_df`:\n{item_ID_values_df}")
            logging.debug(f"JSON with `Item_ID` nesting:\n{item_ID_values_df.to_json(force_ascii=False, indent=4, orient='table', index=True)}")

#ToDo: Create other nested subsections

#Subsection: Add Nested Subsections to Dataframe for Metadata Outside `Attribute_Performance`
outside_attribute_performance_df = outside_attribute_performance_df.set_index(groupby_multiindex)


#Section: Create Final JSON
combined_df = pd.concat([outside_attribute_performance_df, inside_attribute_performance_df], axis='columns', ignore_index=False)
####################
output = combined_df.copy()
purpose = "create-combined-df"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Deduplicate Records
# Using pandas' `duplicated` method on dict data type fields raises a TypeError, so their contents must be compared as equivalent strings
fields_used_in_deduping = [field_name for field_name in fields_used_in_join_multiindex if field_name in combined_df.columns.to_list()]
for field_name in combined_df.columns.to_list():
    if combined_df[field_name].apply(lambda cell_value: repr(type(cell_value))).eq("<class 'dict'>").all():
        string_field_name = f"{field_name}_string"
        combined_df[string_field_name] = combined_df[field_name].astype('string')
        fields_used_in_deduping.append(string_field_name)
combined_df['repeat'] = combined_df.duplicated(subset=fields_used_in_deduping, keep='first')
####################
output = combined_df.copy()
purpose = "combined-df-before-dedupe"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
combined_df = combined_df.loc[combined_df['repeat'] == False]
combined_df = combined_df.drop(columns=fields_used_in_deduping + ['repeat'])
####################
output = combined_df.copy()
purpose = "combined-df-after-dedupe"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.debug(f"`combined_df` after deduping:\n{combined_df}")

#Subsection: Restore Initial Record Order
combined_df = combined_df.reset_index()
combined_df['sort'] = combined_df[groupby_multiindex].apply(
    lambda cell_value: '|'.join(cell_value.astype(str)),  # Combines all values in the fields specified by the index operator of the dataframe to which the `apply` method is applied
    axis='columns',
)
####################
#output = combined_df.copy()
#purpose = "combined-df-add-sort-field"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
record_reordering_dict = {metadata_string: record_sorting_dict[metadata_string] for metadata_string in combined_df['sort'].tolist()}
combined_df['sort'] = combined_df['sort'].replace(record_reordering_dict)
####################
#output = combined_df.copy()
#purpose = "combined-df-with-sort"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
combined_df = combined_df.sort_values(
    by='sort',
    ignore_index=True,  # This resets the record index
)
####################
#output = combined_df.copy()
#purpose = "combined-df-sorted"
#number = number + 1
#output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
#try:
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
#except ValueError:
#    new_index_names = {name:f"_index_{name}" for name in output.index.names}
#    output.index = output.index.set_names(new_index_names)
#    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.info(f"`combined_df` with original record order restored:\n{combined_df}")

#Subsection: Combine All Fields with Proper Nesting
final_df = (combined_df.groupby([f"index_{field_name}" for field_name in metadata_outside_attribute_performance])).apply(lambda x: x[['temp_Attribute_Performance']].to_dict('list')['temp_Attribute_Performance']).rename("Attribute_Performance")
####################
output = final_df.copy()
purpose = "create-final-df"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
final_df = pd.merge(
    outside_attribute_performance_df,
    final_df,  # This argument can be a series, so `final_df` is in the secondary position
    on=[field_name for field_name in groupby_multiindex if field_name[6:] in metadata_outside_attribute_performance],
    suffixes=("_DELETE", None)
)
####################
output = final_df.copy()
purpose = "final-df-all-fields"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
# The merge operation creates a record for each unique combination of all the metadata fields, but a record for each unique combination of the metadata fields outside `Attribute_Performance` is needed at this point; since those metadata fields are the index fields, deduplication by the index is used
final_df['repeat'] = final_df.index.duplicated(keep='first')
final_df = final_df.loc[final_df['repeat'] == False]
final_df = final_df.drop(columns=[field_name for field_name in final_df.columns.to_list() if field_name.endswith("_DELETE")] + fields_to_drop_at_end + ['repeat'])  # Any fields using the values from merge's `suffix` argument are duplicates; this removes one of the duplicates
####################
output = final_df.copy()
purpose = "final-df-pruned"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################

#Subsection: Organize Fields and Data Types
index_field_names = [field_name for field_name in final_df.columns.to_list() if field_name[6:] in metadata_multiindex_fields]
fields_to_drop_at_end = fields_to_drop_at_end + index_field_names
final_df = final_df.drop(columns=fields_to_drop_at_end)
final_df = final_df.replace({"`None`": None})
fields_to_update_dtypes = [field for field in metadata_multiindex_fields if field in final_df.columns.to_list()]
df_dtypes = {key: value for (key, value) in df_dtypes.items() if key in fields_to_update_dtypes}  # JSON uses strings for dates
final_df = final_df.astype(df_dtypes)
logging.info(f"`final_df` summary info:\n{return_string_of_dataframe_info(final_df)}")
####################
output = final_df.copy()
purpose = "final-final-df"
number = number + 1
output.to_csv(directory_with_final_JSONs / f'__{number}_test_{purpose}.csv', encoding='utf-8', errors='backslashreplace')
try:
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
except ValueError:
    new_index_names = {name:f"_index_{name}" for name in output.index.names}
    output.index = output.index.set_names(new_index_names)
    output.to_json(directory_with_final_JSONs / f'__{number}_test_{purpose}.json', force_ascii=False, indent=4, orient='table', index=True)
####################
logging.debug(f"`final_df`:\n{final_df}")
logging.info(f"Final JSON:\n{final_df.to_json(force_ascii=False, indent=4, orient='table', index=False)}")
final_df.to_json(directory_with_final_JSONs / f'{report_name}_final.json', force_ascii=False, indent=4, orient='table', index=False)