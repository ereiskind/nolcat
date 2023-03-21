"""Transform Excel worksheet produced by "tests\\data\\create_JSON_base.json" into a JSON."""

import logging
import pathlib
import io
import re
from openpyxl import load_workbook
import pandas as pd

# `from ..nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat import app` raises `ModuleNotFoundError: No module named 'nolcat'`
# `from .. import nolcat` raises `ImportError: attempted relative import with no known parent package`
# `from .nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat.app import return_string_of_dataframe_info` raises `ModuleNotFoundError: No module named 'nolcat'`

logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s] %(message)s")

absolute_path_to_tests_directory = pathlib.Path(__file__).parent.resolve()
directory_with_final_JSONs = absolute_path_to_tests_directory / 'data' / 'COUNTER_JSONs_for_tests'


def return_string_of_dataframe_info(df):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    in_memory_stream = io.StringIO()
    df.info(buf=in_memory_stream)
    return in_memory_stream.getvalue()


def last_day_of_month(first_day_of_month):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    year_and_month_string = first_day_of_month.date().isoformat()[0:-2]  # Returns an ISO date string, then takes off the last two digits
    return year_and_month_string + str(first_day_of_month.days_in_month)


#Section: Load the Workbook(s)
file_path = input("Enter the complete file path for the Excel workbook output from OpenRefine: ")
file_path = pathlib.Path(file_path)
file = load_workbook(filename=file_path, read_only=True)
report_name = file_path.stem
report_type = report_name.split("_")[1]
sheet = file[report_name]
logging.info(f"Creating JSON from report {report_name} for a {report_type}.")


#Section: Get the Field Names
# Excel workbooks created by Apache POI programs like OpenRefine don't have their `max_column` and `max_row` attributes set properly, so those values must be reset, after which a loop through rows that breaks after the first row can be used
sheet.reset_dimensions()
df_field_names = []
for row in sheet.rows:
    for field_name in row:
        
        # `None` in regex methods raises a TypeError, so they need to be in try-except blocks
        try:
            if re.match(r'^[Cc]omponent', field_name.value):
                continue  # The rarely used `Component` subtype fields aren't captured by this program
        except TypeError:
            pass
        
        if field_name.value is None:
            continue  # Deleted data and merged cells for header values can make Excel think null columns are in use; when read, these columns add `None` to `df_field_names`, causing a `ValueError: Number of passed names did not patch number of header fields in the file` when reading the worksheet contents into a dataframe
        else:
            df_field_names.append(field_name.value)
    break
logging.info(f"The field names are {df_field_names}.")


#Section: Create Dataframe
#Subsection: Ensure String Data Type for Potentially Numeric Metadata Fields
# Strings will be pandas object dtype at this point, but object to string conversion is fairly simple; string fields that pandas might automatically assign a numeric dtype to should be set as strings at the creation of the dataframe to head off problems.
df_dtypes = dict()
for field_name in df_field_names:
    if field_name == "Count" or field_name == "YOP":
        df_dtypes[field_name] = 'Int64'  # The pandas integer dtype; Python's 'int' is ignored
    else:  # JSON uses strings for dates; `Begin_Date` is changed to a date with the `converters` argument in `read_excel`
        df_dtypes[field_name] = 'string'

#Subsection: Create Dataframe from Excel Worksheet
df = pd.read_excel(
    file_path,
    sheet_name=report_name,
    engine='openpyxl',
    header=0,  # This means the first row of the spreadsheet will be treated as the row with the headers, meaning the data starts on the second row
    names=df_field_names,
    dtype=df_dtypes,
    converters={  # `to_datetime` called as object, not function; adding `format` argument causes TypeError
    'Begin_Date': pd.to_datetime,
    'Publication_Date': pd.to_datetime,
    'Parent_Publication_Date': pd.to_datetime,
    }
)
logging.debug(f"Complete dataframe:\n{df}")
logging.info(f"Dataframe summary info:\n{return_string_of_dataframe_info(df)}")


#Section: Update Dataframe
df = df.replace(r'\n', '', regex=True)  # Removes errant newlines found in some reports, primarily at the end of resource names
df = df.replace("licence", "license")  # "Have `license` always use American English spelling
df['End_Date'] = df['Begin_Date'].map(last_day_of_month)
try:
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')
except:
    df['Begin_Date'] = pd.to_datetime(df['Begin_Date'])
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')
if 'YOP' in list(df.columns):
    df['YOP'] = df['YOP'].astype('string')  # Setting the dtype to string when reading in the data means every value in the field ends with `.0`

#Subsection: Put Placeholder in for Null Values
# Null values and fields with nothing but null values cause problems and errors in groupby functions, so they're replaced with placeholder strings
df = df.fillna("`None`")
df = df.replace(
    to_replace='^\s*$',
    # The regex is designed to find the blank but not null cells by finding those cells containing nothing (empty strings) or only whitespace. The whitespace metacharacter `\s` is marked with a deprecation warning, and without the anchors, the replacement is applied not just to whitespaces but to spaces between characters as well.
    value="`None`",
    regex=True
)
logging.debug(f"Dataframe after initial updates:\n{df}")