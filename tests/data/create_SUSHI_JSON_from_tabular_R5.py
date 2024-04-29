"""Transform Excel worksheet produced by "tests\\data\\create_JSON_base.json" into a JSON. This module requires pandas 1.4 or higher to run due to the bug described at https://github.com/pandas-dev/pandas/pull/43949."""

import logging
from pathlib import Path
import io
import re
import json
from openpyxl import load_workbook
import pandas as pd

# `from ..nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat import app` raises `ModuleNotFoundError: No module named 'nolcat'`
# `from .. import nolcat` raises `ImportError: attempted relative import with no known parent package`
# `from .nolcat import app` raises `ImportError: attempted relative import with no known parent package`
# `from nolcat.app import return_string_of_dataframe_info` raises `ModuleNotFoundError: No module named 'nolcat'`

log = logging.getLogger(__name__)

directory_with_final_JSONs = Path(__file__).parent / 'data' / 'COUNTER_JSONs_for_tests'


def return_string_of_dataframe_info(df):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    in_memory_stream = io.StringIO()
    df.info(buf=in_memory_stream)
    return in_memory_stream.getvalue()


def last_day_of_month(first_day_of_month):
    # This is a copy of a function in `nolcat.nolcat.app`, which can't be imported
    year_and_month_string = first_day_of_month.date().isoformat()[0:-2]  # Returns an ISO date string, then takes off the last two digits
    return year_and_month_string + str(first_day_of_month.days_in_month)


def correct_item_parent_dictionary(item_parent_dictionary):
    """This function corrects the problems in the nested `Item_Parent` JSON that come from the dataframe to dictionary/JSON conversion process.

    There's no way to directly create the `Item_Parent` nested JSON section from its subsections and the existing columns; this function, entered into a lambda statement, performs the final adjustments.

    Args:
        item_parent_dictionary (dict): an individual value in the series created by applying a groupby operation to `item_parent_values_df`
    
    Returns:
        dict: the dictionary that's the value to the `Item_Parent` key in the SUSHI test data JSON
    """
    for metadata in item_parent_dictionary.values():  # This removes the key with the record number
        corrected_metadata = dict()  # This creates an empty dictionary to avoid changing `metadata` while iterating through it
        for key, value in metadata.items():
            if isinstance(value, list):  # Key-value pairs with list and string (non-list) values are separated because using the comprehensive null checking method on a list raises an error
                if pd.isnull(value).all():
                    continue  # Key-value pair not being added to `corrected_metadata`
                else:
                    corrected_metadata[key] = value
            else:
                if pd.isnull(value):
                    continue  # Key-value pair not being added to `corrected_metadata`
                elif key == "Parent_Title":
                    corrected_metadata['Item_Name'] = value
                elif key == "Parent_Data_Type":
                    corrected_metadata['Data_Type'] = value
                else:
                    corrected_metadata[key] = value
        return corrected_metadata


#Section: Load the Workbook(s)
file_path = input("Enter the complete file path for the Excel workbook output from OpenRefine: ")
file_path = Path(file_path)
file = load_workbook(filename=file_path, read_only=True)
report_name = file_path.stem
report_type = report_name.split("_")[1]
sheet = file[report_name]
log.info(f"Creating JSON from report {report_name} for a {report_type}.")


#Section: Get the Field Names
# Excel workbooks created by Apache POI programs like OpenRefine don't have their `max_column` and `max_row` attributes set properly, so those values must be reset, after which a loop through rows that breaks after the first row can be used
sheet.reset_dimensions()
df_field_names = []
for row in sheet.rows:
    for field_name in row:
        
        # `None` in regex methods raises a TypeError, so they need to be in try-except blocks
        try:
            if re.fullmatch(r"^[Cc]omponent", field_name.value):
                continue  # The rarely used `Component` subtype fields aren't captured by this program
        except TypeError:
            pass
        
        if field_name.value is None:
            continue  # Deleted data and merged cells for header values can make Excel think null columns are in use; when read, these columns add `None` to `df_field_names`, causing a `ValueError: Number of passed names did not patch number of header fields in the file` when reading the worksheet contents into a dataframe
        else:
            df_field_names.append(field_name.value)
    break
log.info(f"The field names are {df_field_names}.")


#Section: Create Dataframe
#Subsection: Ensure String Data Type for Potentially Numeric Metadata Fields
# Strings will be pandas object dtype at this point, but object to string conversion is fairly simple; string fields that pandas might automatically assign a numeric dtype to should be set as strings at the creation of the dataframe to head off problems.
df_dtypes = dict()
for field_name in df_field_names:
    if field_name == "Count" or field_name == "YOP":
        df_dtypes[field_name] = 'Int64'  # The pandas integer dtype; Python's 'int' is ignored
    else:  # JSON uses strings for dates; `Begin_Date` is changed to a date with the `converters` argument in `read_excel`
        df_dtypes[field_name] = 'string'

#Subsection: Create Dataframe from Excel Worksheet
df = pd.read_excel(
    file_path,
    sheet_name=report_name,
    engine='openpyxl',
    header=0,  # This means the first row of the spreadsheet will be treated as the row with the headers, meaning the data starts on the second row
    names=df_field_names,
    dtype=df_dtypes,
    converters={  # `to_datetime` called as object, not function; adding `format` argument causes TypeError
    'Begin_Date': pd.to_datetime,
    'Publication_Date': pd.to_datetime,
    'Parent_Publication_Date': pd.to_datetime,
    }
)
log.debug(f"Complete dataframe:\n{df}")
log.info(f"Dataframe summary info:\n{return_string_of_dataframe_info(df)}")


#Section: Update Dataframe
df = df.replace(r"\n", "", regex=True)  # Removes errant newlines found in some reports, primarily at the end of resource names
df = df.replace("licence", "license")  # "Have `license` always use American English spelling
df['End_Date'] = df['Begin_Date'].map(last_day_of_month)
try:
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')
except:
    df['Begin_Date'] = pd.to_datetime(df['Begin_Date'])
    df['Begin_Date'] = df['Begin_Date'].dt.strftime('%Y-%m-%d')

#Subsection: Put Placeholder in for Null Values
# Null values and fields with nothing but null values cause problems and errors in groupby functions, so they're replaced with placeholder strings
df = df.fillna("`None`")
df = df.replace(
    to_replace='^\s*$',
    # The regex is designed to find the blank but not null cells by finding those cells containing nothing (empty strings) or only whitespace. The whitespace metacharacter `\s` is marked with a deprecation warning, and without the anchors, the replacement is applied not just to whitespaces but to spaces between characters as well.
    value="`None`",
    regex=True
)
log.debug(f"Dataframe after initial updates:\n{df}")


######Section: CHANGE DATAFRAME INTO JSON #####

#Section: Create Field Lists and Dataframes for Multiindexes, Groupby Operations, and Dataframe Recombination
fields_used_in_performance_nested_groups = ['Begin_Date', 'End_Date', 'Metric_Type', 'Count']
fields_used_for_groupby_operations = [field_name for field_name in df_field_names if field_name not in fields_used_in_performance_nested_groups]

fields_used_in_performance_join_multiindex = fields_used_for_groupby_operations + ['Begin_Date']
performance_join_multiindex_df = df[fields_used_in_performance_join_multiindex].set_index(fields_used_for_groupby_operations, drop=False)
fields_to_drop_at_end = []
possible_fields_in_item_ID = ['DOI', 'Proprietary_ID', 'ISBN', 'Print_ISSN', 'Online_ISSN', 'URI']  # List defined here because it's used in two separate `if` blocks

# The default text sorting methods in OpenRefine and Pandas order non-letter characters differently, meaning applying a pandas sort to the resource names won't put them in the same order as they were in openRefine after the sort; to return the records to their same order at the end of the module, their order must be saved at this point
record_sorting_strings = df[fields_used_for_groupby_operations].apply(
    lambda cell_value: '|'.join(cell_value.astype(str)),  # Combines all values in the fields specified by the index operator of the dataframe to which the `apply` method is applied
    axis='columns',
)
record_sorting_strings = record_sorting_strings.drop_duplicates().reset_index(drop=True)
log.debug(f"Complete metadata in order from OpenRefine:\n{record_sorting_strings}")
record_sorting_dict = record_sorting_strings.to_dict()
record_sorting_dict = {metadata_string: order_number for (order_number, metadata_string) in record_sorting_dict.items()}
log.info(f"Metadata strings with ordering numbers:\n{record_sorting_dict}")

#Section: Create Nested JSON Section for Publisher IDs
if report_type == "DR" or report_type == "TR" or report_type == "IR":
    if 'Publisher_ID' in list(performance_join_multiindex_df.columns):  # If the publisher ID field exists
        fields_to_drop_at_end.append('Publisher_ID')
        if not performance_join_multiindex_df['Publisher_ID'].eq("`None`").all():  # If the publisher ID field has values
            publisher_ID_values_df = performance_join_multiindex_df.copy()
            non_publisher_ID_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name != "Publisher_ID"]
            publisher_ID_values_df = publisher_ID_values_df.drop(columns=non_publisher_ID_fields)
            publisher_ID_values_df['Type'] = "Proprietary"
            publisher_ID_values_df = publisher_ID_values_df.rename(columns={"Publisher_ID": "Value"})
            publisher_ID_values_df = publisher_ID_values_df.replace({"`None`": None})
            publisher_ID_values_df = publisher_ID_values_df.reset_index()
            publisher_ID_values_df['repeat'] = publisher_ID_values_df.duplicated(subset=fields_used_for_groupby_operations, keep='first')
            publisher_ID_values_df =  publisher_ID_values_df.loc[publisher_ID_values_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
            publisher_ID_values_df =  publisher_ID_values_df.drop(columns=['repeat'])
            publisher_ID_values_df = (publisher_ID_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda publisher_ID_groupby: publisher_ID_groupby[['Type', 'Value']].to_dict('records')).rename("Publisher_ID")
            log.debug(f"`publisher_ID_values_df`:\n{publisher_ID_values_df}")
            # `to_json` method raises `ValueError: Overlapping names between the index and columns` because `Publisher_ID` is both the key to the list of dictionaries just formed and a field in the multiindex


#Section: Create Nested JSON Section for Item IDs
#Subsection: Create Nested JSON Section for DR Item IDs
if report_type == "DR":
    if 'Proprietary_ID' in list(performance_join_multiindex_df.columns):  # If the proprietary ID field, the only item ID field in DR, exists
        fields_to_drop_at_end.append('Proprietary_ID')
        if not performance_join_multiindex_df['Proprietary_ID'].eq("`None`").all():  # If the proprietary ID field has values
            item_ID_values_df = performance_join_multiindex_df.copy()
            non_item_ID_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name != "Proprietary_ID"]
            item_ID_values_df = item_ID_values_df.drop(columns=non_item_ID_fields)
            item_ID_values_df['Type'] = "Proprietary"
            item_ID_values_df = item_ID_values_df.rename(columns={"Proprietary_ID": "Value"})
            item_ID_values_df = item_ID_values_df.replace({"`None`": None})
            item_ID_values_df = item_ID_values_df.reset_index()
            item_ID_values_df['repeat'] = item_ID_values_df.duplicated(subset=fields_used_for_groupby_operations, keep='first')
            item_ID_values_df = item_ID_values_df.loc[item_ID_values_df['repeat'] == False]
            item_ID_values_df = item_ID_values_df.drop(columns=['repeat'])
            item_ID_values_df = (item_ID_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_ID_groupby: item_ID_groupby[['Type', 'Value']].to_dict('records')).rename("Item_ID")
            log.debug(f"`item_ID_values_df`:\n{item_ID_values_df}")
            log.debug(f"JSON with `Item_ID` nesting:\n{item_ID_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Subsection: Create Nested JSON Section for TR and IR Item IDs
if report_type == "TR" or report_type == "IR":
    if 'DOI' in list(performance_join_multiindex_df.columns) or 'Proprietary_ID' in list(performance_join_multiindex_df.columns) or 'ISBN' in list(performance_join_multiindex_df.columns) or 'Print_ISSN' in list(performance_join_multiindex_df.columns) or 'Online_ISSN' in list(performance_join_multiindex_df.columns) or 'URI' in list(performance_join_multiindex_df.columns):  # If the fields in the item ID section exist
        if not performance_join_multiindex_df['DOI'].eq("`None`").all() or not performance_join_multiindex_df['Proprietary_ID'].eq("`None`").all() or not performance_join_multiindex_df['ISBN'].eq("`None`").all() or not performance_join_multiindex_df['Print_ISSN'].eq("`None`").all() or not performance_join_multiindex_df['Online_ISSN'].eq("`None`").all() or not performance_join_multiindex_df['URI'].eq("`None`").all():  # If the fields in the item ID section have values
            item_ID_values_df = performance_join_multiindex_df.copy()

            #Subsection: Determine All the Fields Going in the Nested Section
            fields_in_item_ID = []
            for field_name in item_ID_values_df.columns:
                if field_name in possible_fields_in_item_ID:
                    fields_to_drop_at_end.append(field_name)
                    if not item_ID_values_df[field_name].eq("`None`").all():
                        fields_in_item_ID.append(field_name)

            #Subsection: Remove Fields Not Being Nested
            non_item_ID_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name not in fields_in_item_ID]
            item_ID_values_df = item_ID_values_df.drop(columns=non_item_ID_fields).drop(columns=['Begin_Date'])
            item_ID_values_df = item_ID_values_df.stack().reset_index()  # If the index isn't reset, the stack method returns a series
            item_ID_values_df = item_ID_values_df.rename(columns={item_ID_values_df.columns[-2]: 'Type', 0: 'Value'})  # The name of the `Type` field is `level_#` where `#` is the position in a zero-based order of the columns in the dataframe; since the exact name that needs to be changed cannot be know in advanced, it must be found from its penultimate position in the list of field names

            #Subsection: Remove Null Values and Repetitions
            item_ID_values_df = item_ID_values_df.loc[item_ID_values_df['Value'] != "`None`"]  # Because the null placeholders are targeted for removal, they don't need to be replaced
            item_ID_values_df['repeat'] = item_ID_values_df.duplicated(keep='first')
            item_ID_values_df = item_ID_values_df.loc[item_ID_values_df['repeat'] == False]
            item_ID_values_df = item_ID_values_df.drop(columns=['repeat'])

            #Subsection: Complete Nested JSON Section for Item IDs
            item_ID_values_df = (item_ID_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_ID_groupby: item_ID_groupby[['Type', 'Value']].to_dict('records')).rename("Item_ID")
            log.debug(f"`item_ID_values_df`:\n{item_ID_values_df}")
            log.debug(f"JSON with `Item_ID` nesting:\n{item_ID_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Nested JSON Section for Authors
if report_type == "IR":
    if 'Authors' in list(performance_join_multiindex_df.columns):  # If the author field exists
        fields_to_drop_at_end.append('Authors')
        if not performance_join_multiindex_df['Authors'].eq("`None`").all():  # If the author field has values
            author_values_df = performance_join_multiindex_df.copy()
            non_author_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name != "Authors"]
            author_values_df = author_values_df.drop(columns=non_author_fields)
            author_values_df['Type'] = "Author"
            author_values_df = author_values_df.rename(columns={"Authors": "Name"})
            author_values_df = author_values_df.replace({"`None`": None})
            author_values_df = author_values_df.reset_index()
            author_values_df['repeat'] = author_values_df.duplicated(subset=fields_used_for_groupby_operations, keep='first')
            author_values_df =  author_values_df.loc[author_values_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
            author_values_df =  author_values_df.drop(columns=['repeat'])
            author_values_df = (author_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda authors_groupby: authors_groupby[['Type', 'Name']].to_dict('records')).rename("Item_Contributors")  # `Item_Contributors` uses `Type` and `Name` as the keys in its dictionaries
            log.debug(f"`author_values_df`:\n{author_values_df}")
            log.debug(f"JSON with `Item_Contributors` nesting:\n{author_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Nested JSON Section for Publication Date
if report_type == "IR":
    if 'Publication_Date' in list(performance_join_multiindex_df.columns):  # If the publication date field exists
        fields_to_drop_at_end.append('Publication_Date')
        if not performance_join_multiindex_df['Publication_Date'].eq("`None`").all():  # If the publication date field has values
            publication_date_values_df = performance_join_multiindex_df.copy()
            # For the `.dt` accessor function that removes the time data to work, the field must have a date data type, meaning there are no strings in the field
            publication_date_values_df['Publication_Date'] = publication_date_values_df['Publication_Date'].replace(
                to_replace="`None`",
                value=pd.NA,
            )
            publication_date_values_df['Publication_Date'] = pd.to_datetime(publication_date_values_df['Publication_Date'])
            publication_date_values_df['Publication_Date'] = publication_date_values_df['Publication_Date'].dt.strftime('%Y-%m-%d')
            publication_date_values_df['Publication_Date'] = publication_date_values_df['Publication_Date'].fillna("`None`")
            non_publication_date_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name != "Publication_Date"]
            publication_date_values_df = publication_date_values_df.drop(columns=non_publication_date_fields)
            publication_date_values_df['Type'] = "Publication_Date"
            publication_date_values_df = publication_date_values_df.rename(columns={"Publication_Date": "Value"})
            publication_date_values_df = publication_date_values_df.replace({"`None`": None})
            publication_date_values_df = publication_date_values_df.reset_index()
            publication_date_values_df['repeat'] = publication_date_values_df.duplicated(subset=fields_used_for_groupby_operations, keep='first')
            publication_date_values_df =  publication_date_values_df.loc[publication_date_values_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
            publication_date_values_df =  publication_date_values_df.drop(columns=['repeat'])
            publication_date_values_df = (publication_date_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_dates_groupby: item_dates_groupby[['Type', 'Value']].to_dict('records')).rename("Item_Dates")
            log.debug(f"`publication_date_values_df`:\n{publication_date_values_df}")
            log.debug(f"JSON with `Item_Dates` nesting:\n{publication_date_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Nested JSON Section for Article Version
if report_type == "IR":
    if 'Article_Version' in list(performance_join_multiindex_df.columns):  # If the article version field exists
        fields_to_drop_at_end.append('Article_Version')
        if not performance_join_multiindex_df['Article_Version'].eq("`None`").all():  # If the article version field has values
            article_version_values_df = performance_join_multiindex_df.copy()
            non_article_version_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name != "Article_Version"]
            article_version_values_df = article_version_values_df.drop(columns=non_article_version_fields)
            article_version_values_df['Type'] = "Article_Version"
            article_version_values_df = article_version_values_df.rename(columns={"Article_Version": "Value"})
            article_version_values_df = article_version_values_df.replace({"`None`": None})
            article_version_values_df = article_version_values_df.reset_index()
            article_version_values_df['repeat'] = article_version_values_df.duplicated(subset=fields_used_for_groupby_operations, keep='first')
            article_version_values_df =  article_version_values_df.loc[article_version_values_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
            article_version_values_df =  article_version_values_df.drop(columns=['repeat'])
            article_version_values_df = (article_version_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_attributes_groupby: item_attributes_groupby[['Type', 'Value']].to_dict('records')).rename("Item_Attributes")
            log.debug(f"`article_version_values_df`:\n{article_version_values_df}")
            log.debug(f"JSON with `Item_Attributes` nesting containing `Article_Version`:\n{article_version_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Nested JSON Section for Item Parent Data
# Parent data fields and their labels with nesting in the JSON
    # Parent_Title --> Item_Parent > Item_Name
    # Parent_Authors --> Item_Parent > Item_Contributors > Name
    # Parent_Publication_Date --> Item_Parent > Item_Dates > Value
    # Parent_Article_Version --> Item_Parent > Item_Attributes > Value
    # Parent_Data_Type --> Item_Parent > Data_Type
    # Parent_DOI --> Item_Parent > Item_ID > Value
    # Parent_Proprietary_ID --> Item_Parent > Item_ID > Value
    # Parent_ISBN --> Item_Parent > Item_ID > Value
    # Parent_Print_ISSN --> Item_Parent > Item_ID > Value
    # Parent_Online_ISSN --> Item_Parent > Item_ID > Value
    # Parent_URI --> Item_Parent > Item_ID > Value
if report_type == "IR":
    if 'Parent_Title' in list(performance_join_multiindex_df.columns) or 'Parent_Authors' in list(performance_join_multiindex_df.columns) or 'Parent_Publication_Date' in list(performance_join_multiindex_df.columns) or 'Parent_Article_Version' in list(performance_join_multiindex_df.columns) or 'Parent_Data_Type' in list(performance_join_multiindex_df.columns) or 'Parent_DOI' in list(performance_join_multiindex_df.columns) or 'Parent_Proprietary_ID' in list(performance_join_multiindex_df.columns) or 'Parent_ISBN' in list(performance_join_multiindex_df.columns) or 'Parent_Print_ISSN' in list(performance_join_multiindex_df.columns) or 'Parent_Online_ISSN' in list(performance_join_multiindex_df.columns) or 'Parent_URI' in list(performance_join_multiindex_df.columns):  # If the fields in the item parent section exist
        if not performance_join_multiindex_df['Parent_Title'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Authors'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Publication_Date'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Article_Version'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Data_Type'].eq("`None`").all() or not performance_join_multiindex_df['Parent_DOI'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Proprietary_ID'].eq("`None`").all() or not performance_join_multiindex_df['Parent_ISBN'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Print_ISSN'].eq("`None`").all() or not performance_join_multiindex_df['Parent_Online_ISSN'].eq("`None`").all() or not performance_join_multiindex_df['Parent_URI'].eq("`None`").all():  # If the fields in the item parent section have values
            item_parent_values_df = performance_join_multiindex_df.copy()

            #Subsection: Determine All the Fields Going in the Nested Section
            possible_fields_in_item_parent = ['Parent_Title', 'Parent_Authors', 'Parent_Publication_Date', 'Parent_Article_Version', 'Parent_Data_Type', 'Parent_DOI', 'Parent_Proprietary_ID', 'Parent_ISBN', 'Parent_Print_ISSN', 'Parent_Online_ISSN', 'Parent_URI']
            fields_in_item_parent = []
            for field_name in item_parent_values_df.columns:
                if field_name in possible_fields_in_item_parent:
                    fields_to_drop_at_end.append(field_name)
                    if not item_parent_values_df[field_name].eq("`None`").all():
                        fields_in_item_parent.append(field_name)
            if 'Parent_Publication_Date' in fields_in_item_parent:
                # For the `.dt` accessor function that removes the time data to work, the field must have a date data type, meaning there are no strings in the field
                item_parent_values_df['Parent_Publication_Date'] = item_parent_values_df['Parent_Publication_Date'].replace(
                    to_replace="`None`",
                    value=pd.NA,
                )
                item_parent_values_df['Parent_Publication_Date'] = pd.to_datetime(item_parent_values_df['Parent_Publication_Date'])
                item_parent_values_df['Parent_Publication_Date'] = item_parent_values_df['Parent_Publication_Date'].dt.strftime('%Y-%m-%d')
                item_parent_values_df['Parent_Publication_Date'] = item_parent_values_df['Parent_Publication_Date'].fillna("`None`")
            
            #Subsection: Remove Fields Not Being Nested
            non_item_parent_fields = [field_name for field_name in fields_used_for_groupby_operations if field_name not in fields_in_item_parent]
            item_parent_values_df = item_parent_values_df.drop(columns=non_item_parent_fields).drop(columns=['Begin_Date'])

            #Subsection: Prepare for Inner Nesting
            item_parent_values_df = item_parent_values_df.stack().reset_index()  # If the index isn't reset, the stack method returns a series
            item_parent_values_df = item_parent_values_df.rename(columns={item_parent_values_df.columns[-2]: 'Type', 0: 'Value'})  # The name of the `Type` field is `level_#` where `#` is the position in a zero-based order of the columns in the dataframe; since the exact name that needs to be changed cannot be know in advanced, it must be found from its penultimate position in the list of field names
            item_parent_values_df['Type'] = item_parent_values_df['Type'].map(lambda type: re.search(r"Parent_(.+)", type)[1] if isinstance(type, str) else type).replace("Title", "Item_Name").replace("Authors", "Author")

            #Subsection: Remove Null Values and Repetitions
            item_parent_values_df = item_parent_values_df.loc[item_parent_values_df['Value'] != "`None`"]  # Because the null placeholders are targeted for removal, they don't need to be replaced
            item_parent_values_df['repeat'] = item_parent_values_df.duplicated(keep='first')
            item_parent_values_df = item_parent_values_df.loc[item_parent_values_df['repeat'] == False]
            item_parent_values_df = item_parent_values_df.drop(columns=['repeat'])
            item_parent_values_df = item_parent_values_df.set_index(fields_used_for_groupby_operations)

            #Subsection: Create Inner Nested JSON Section for Author
            item_contributors_df = item_parent_values_df.loc[item_parent_values_df['Type'] == 'Author']
            if not item_contributors_df.empty:
                item_contributors_df = item_contributors_df.rename(columns={'Value': 'Name'})
                item_contributors_df = (item_contributors_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_contributors_groupby: item_contributors_groupby[['Type', 'Name']].to_dict('records')).rename("Item_Contributors")
                item_parent_values_df = item_parent_values_df.join(item_contributors_df)
                log.debug(f"`item_parent_values_df`:\n{item_parent_values_df}")
                log.debug(f"JSON with `Item_Contributors` nesting:\n{item_parent_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

            #Subsection: Create Inner Nested JSON Section for Publication Date
            item_dates_df = item_parent_values_df.loc[item_parent_values_df['Type'] == 'Publication_Date']
            if not item_dates_df.empty:
                item_dates_df = (item_dates_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_dates_groupby: item_dates_groupby[['Type', 'Value']].to_dict('records')).rename("Item_Dates")
                item_parent_values_df = item_parent_values_df.join(item_dates_df)
                log.debug(f"`item_parent_values_df`:\n{item_parent_values_df}")
                log.debug(f"JSON with `Item_Dates` nesting:\n{item_parent_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

            #Subsection: Create Inner Nested JSON Section for Article Version
            item_attributes_df = item_parent_values_df.loc[item_parent_values_df['Type'] == 'Article_Version']
            if not item_attributes_df.empty:
                # NoLCAT test data never enters this code block
                item_attributes_df = (item_attributes_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_attributes_groupby: item_attributes_groupby[['Type', 'Value']].to_dict('records')).rename("Item_Attributes")
                item_parent_values_df = item_parent_values_df.join(item_attributes_df)
                log.debug(f"`item_parent_values_df`:\n{item_parent_values_df}")
                log.debug(f"JSON with `Item_Attributes` nesting:\n{item_parent_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

            #Subsection: Create Inner Nested JSON Section for Item IDs
            item_ID_df = item_parent_values_df[item_parent_values_df['Type'].isin(possible_fields_in_item_ID)]
            if not item_ID_df.empty:
                item_ID_df = (item_ID_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_ID_groupby: item_ID_groupby[['Type', 'Value']].to_dict('records')).rename("Item_ID")
                item_parent_values_df = item_parent_values_df.join(item_ID_df)
                log.debug(f"`item_parent_values_df`:\n{item_parent_values_df}")
                log.debug(f"JSON with `Item_ID` nesting:\n{item_parent_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

            #Subsection: Deduplicate Dataframe with All Inner Nested Sections
            item_parent_values_df = item_parent_values_df.reset_index()
            item_parent_subsection_string_fields = []
            item_parent_fields_to_nest = []
            # Using pandas' `duplicated` method on dict data type fields raises a TypeError, so their contents must be compared as equivalent strings
            if 'Item_Contributors' in list(item_parent_values_df.columns):
                item_parent_values_df['item_contributors_string'] = item_parent_values_df['Item_Contributors'].astype('string')
                item_parent_subsection_string_fields.append('item_contributors_string')
                item_parent_fields_to_nest.append('Item_Contributors')
            if 'Item_Dates' in list(item_parent_values_df.columns):
                item_parent_values_df['item_dates_string'] = item_parent_values_df['Item_Dates'].astype('string')
                item_parent_subsection_string_fields.append('item_dates_string')
                item_parent_fields_to_nest.append('Item_Dates')
            if 'Item_Attributes' in list(item_parent_values_df.columns):
                item_parent_values_df['item_attributes_string'] = item_parent_values_df['Item_Attributes'].astype('string')
                item_parent_subsection_string_fields.append('item_attributes_string')
                item_parent_fields_to_nest.append('Item_Attributes')
            if 'Item_ID' in list(item_parent_values_df.columns):
                item_parent_values_df['item_ID_string'] = item_parent_values_df['Item_ID'].astype('string')
                item_parent_subsection_string_fields.append('item_ID_string')
                item_parent_fields_to_nest.append('Item_ID')
            item_parent_values_df['repeat'] = item_parent_values_df.duplicated(subset=item_parent_subsection_string_fields + fields_used_for_groupby_operations, keep='first')
            item_parent_values_df = item_parent_values_df.loc[item_parent_values_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
            item_parent_values_df = item_parent_values_df.drop(columns=item_parent_subsection_string_fields).drop(columns=['Type', 'Value', 'repeat'])

            #Subsection: Combine All Inner Nested Sections
            item_parent_values_df = item_parent_values_df.reset_index(drop=True)
            if 'Parent_Title' in list(item_parent_values_df.columns):
                item_parent_fields_to_nest.append('Parent_Title')
            if 'Parent_Data_Type' in list(item_parent_values_df.columns):
                item_parent_fields_to_nest.append('Parent_Data_Type')
            item_parent_values_df = (item_parent_values_df.groupby(fields_used_for_groupby_operations)).apply(lambda item_parent_groupby: item_parent_groupby[item_parent_fields_to_nest].to_dict('index')).apply(correct_item_parent_dictionary).rename("Item_Parent")
            log.debug(f"`item_parent_values_df`:\n{item_parent_values_df}")
            log.debug(f"JSON with `Item_Parent` nesting:\n{item_parent_values_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Nested JSON Section for Performance
#Subsection: Create Instance Grouping
instance_groupby_operation_fields = fields_used_for_groupby_operations + ['Begin_Date']
instance_df = (df.groupby(instance_groupby_operation_fields)).apply(lambda instance_groupby: instance_groupby[['Metric_Type', 'Count']].to_dict('records')).reset_index().rename(columns={0: "Instance"})  # `instance_groupby_operation_fields` contains all the non-null fields that must be the same for all the records represented in a instance grouping plus a date field to keep instances where the metric and count are repeated from being combined
instance_df = instance_df.set_index(fields_used_in_performance_join_multiindex)
log.debug(f"`instance_df`:\n{instance_df}")
log.debug(f"JSON with `Instance` nesting:\n{instance_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

#Subsection: Create Period Grouping
df['temp'] = range(1, len(df.index)+1)  # The way groupby works, for each resource defined by metadata, if a given instance (a metric and its count) occurs in multiple months, those months will be combined, which is not the desired behavior; this field puts a unique value in each row/record, which prevents this grouping
period_groupby_operation_fields = fields_used_for_groupby_operations + ['Metric_Type', 'Count', 'temp']
period_df = (df.groupby(period_groupby_operation_fields)).apply(lambda period_groupby: period_groupby[['Begin_Date','End_Date']].to_dict('records')).reset_index().rename(columns={0: "Period"})
period_df = period_df.drop(columns=['temp', 'Metric_Type', 'Count'])
period_df['Period'] = period_df['Period'].map(lambda list_like: list_like[0])
period_df['Begin_Date'] = period_df['Period'].astype('string').map(lambda string: string[16:-28])  # This turns the JSON/dict value into a string, then isolates the desired date from within each string
period_df = period_df.set_index(fields_used_in_performance_join_multiindex)


#Section: Combine Nested JSON Groupings
#Subsection: Combine Period and Instance Groupings
combined_df = period_df.join(instance_df, how='inner')  # This contains duplicate records
combined_df = combined_df.reset_index()
combined_df = combined_df.drop(columns=['Begin_Date'])

#Subsection: Deduplicate Records
# Using pandas' `duplicated` method on dict data type fields raises a TypeError, so their contents must be compared as equivalent strings
combined_df['instance_string'] = combined_df['Instance'].astype('string')
combined_df['period_string'] = combined_df['Period'].astype('string')
fields_to_use_in_deduping = fields_used_for_groupby_operations + ['instance_string'] + ['period_string']
combined_df['repeat'] = combined_df.duplicated(subset=fields_to_use_in_deduping, keep='first')
combined_df = combined_df.loc[combined_df['repeat'] == False]  # Where the Boolean indicates if the record is the same as an earlier record
combined_df = combined_df.drop(columns=['instance_string', 'period_string', 'repeat'])
log.debug(f"`combined_df`:\n{combined_df}")
log.debug(f"JSON with `Period` nesting:\n{combined_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

#Subsection: Create Performance Grouping
combined_df = combined_df.reset_index(drop=True)
joining_df = (combined_df.groupby(fields_used_for_groupby_operations)).apply(lambda performance_groupby: performance_groupby[['Period', 'Instance']].to_dict('records')).reset_index().rename(columns={0: "Performance"})
log.debug(f"`joining_df` before metadata additions:\n{joining_df}")
log.debug(f"JSON with `Performance` nesting:\n{joining_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")

#Subsection: Add Other JSON Groupings
joining_df = joining_df.set_index(fields_used_for_groupby_operations)
if dict(locals()).get('publisher_ID_values_df') is not None:
    joining_df = joining_df.join(publisher_ID_values_df, how='left')
    joining_df = joining_df.rename(columns={'Publisher_ID': '_Publisher_ID'})  # This field and its origin field have the same name, which causes an error when resetting the index; to avoid it, a temporary new name is used
if dict(locals()).get('item_ID_values_df') is not None:
    joining_df = joining_df.join(item_ID_values_df, how='left')
if dict(locals()).get('author_values_df') is not None:
    joining_df = joining_df.join(author_values_df, how='left')
if dict(locals()).get('publication_date_values_df') is not None:
    joining_df = joining_df.join(publication_date_values_df, how='left')
if dict(locals()).get('article_version_values_df') is not None:
    joining_df = joining_df.join(article_version_values_df, how='left')
if dict(locals()).get('item_parent_values_df') is not None:
    joining_df = joining_df.join(item_parent_values_df, how='left')
log.debug(f"`joining_df` after metadata additions:\n{joining_df}")
log.debug(f"JSON with all nesting combined:\n{joining_df.head(10).to_json(force_ascii=False, indent=4, orient='table', index=True)}")


#Section: Create Final JSON
#Subsection: Restore Initial Record Order
final_df = joining_df.reset_index()
final_df['sort'] = final_df[fields_used_for_groupby_operations].apply(
    lambda cell_value: '|'.join(cell_value.astype(str)),  # Combines all values in the fields specified by the index operator of the dataframe to which the `apply` method is applied
    axis='columns',
)
fields_to_drop_at_end.append('sort')  # So this field is removed in the next subsection
record_reordering_dict = {metadata_string: record_sorting_dict[metadata_string] for metadata_string in final_df['sort'].tolist()}
final_df['sort'] = final_df['sort'].replace(record_reordering_dict)
final_df = final_df.sort_values(
    by='sort',
    ignore_index=True,  # This resets the record index
)
log.info(f"`final_df` with original record order restored:\n{final_df}")

#Subsection: Organize Fields and Data Types
final_df = final_df.drop(columns=fields_to_drop_at_end)
if '_Publisher_ID' in list(final_df.columns):
    final_df = final_df.rename(columns={'_Publisher_ID': 'Publisher_ID'})
final_df = final_df.replace({"`None`": None})
fields_to_update_dtypes = [field for field in fields_used_for_groupby_operations if field not in fields_to_drop_at_end]  # Fields in `final_df` not used in groupby operations are formatted as JSONs to become the nested sections of the final JSON, changing their dtypes could interfere with this
df_dtypes = {key: value for (key, value) in df_dtypes.items() if key in fields_to_update_dtypes}  # JSON uses strings for dates
final_df = final_df.astype(df_dtypes)
log.info(f"`final_df` summary info:\n{return_string_of_dataframe_info(final_df)}")
log.debug(f"`final_df`:\n{final_df}")
log.info(f"Final JSON:\n{final_df.to_json(force_ascii=False, indent=4, orient='table', index=False)}")
final_df.to_json(directory_with_final_JSONs / f'{report_name}_final.json', force_ascii=False, indent=4, orient='table', index=False)